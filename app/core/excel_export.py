
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import Project

CHECK_MARK = "\u2713"  # ✓

# --- style constants -----------------------------------------------------
REGION_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
VARIANT_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
PARTS_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

REGION_FONT = Font(bold=True, color="FFFFFF", size=11)
VARIANT_FONT = Font(bold=True, color="1F1F1F", size=10)
PARTS_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PART_ID_FONT = Font(bold=False, size=10)
CHECK_FONT = Font(bold=True, size=11, color="1F7A3D")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT_CENTER = Alignment(horizontal="left", vertical="center")

THIN = Side(style="thin", color="B7B7B7")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def export_to_excel(project: Project, path: str | Path) -> None:
    
    columns = project.all_columns()  # ordered list of ColumnKey, drives everything below

    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = "Data Configuration"

    PARTS_COL = 1
    HEADER_ROW_REGION = 1
    HEADER_ROW_VARIANT = 2
    FIRST_DATA_ROW = 3

    # --- "Parts" header, merged vertically across both header rows ---
    ws.merge_cells(
        start_row=HEADER_ROW_REGION, start_column=PARTS_COL,
        end_row=HEADER_ROW_VARIANT, end_column=PARTS_COL,
    )
    parts_header_cell = ws.cell(row=HEADER_ROW_REGION, column=PARTS_COL, value="Parts")
    parts_header_cell.font = PARTS_HEADER_FONT
    parts_header_cell.fill = PARTS_HEADER_FILL
    parts_header_cell.alignment = CENTER
    parts_header_cell.border = CELL_BORDER
    # Border the merged-but-empty bottom half of the merge too
    ws.cell(row=HEADER_ROW_VARIANT, column=PARTS_COL).border = CELL_BORDER

    # --- Region (row 1, merged) + Variant (row 2) headers, dynamic width ---
    col_index = PARTS_COL + 1
    for region in project.regions:
        span = len(region.variants)
        if span == 0:
            continue  # validation should prevent this, but stay defensive

        start_col = col_index
        end_col = col_index + span - 1

        ws.merge_cells(
            start_row=HEADER_ROW_REGION, start_column=start_col,
            end_row=HEADER_ROW_REGION, end_column=end_col,
        )
        region_cell = ws.cell(row=HEADER_ROW_REGION, column=start_col, value=region.name)
        region_cell.font = REGION_FONT
        region_cell.fill = REGION_FILL
        region_cell.alignment = CENTER

        for c in range(start_col, end_col + 1):
            ws.cell(row=HEADER_ROW_REGION, column=c).border = CELL_BORDER

        for i, variant in enumerate(region.variants):
            vcol = start_col + i
            vcell = ws.cell(row=HEADER_ROW_VARIANT, column=vcol, value=variant)
            vcell.font = VARIANT_FONT
            vcell.fill = VARIANT_FILL
            vcell.alignment = CENTER
            vcell.border = CELL_BORDER

        col_index = end_col + 1

    last_col = col_index - 1  # last column used by the matrix

    # --- Body: one row per part ---
    row = FIRST_DATA_ROW
    for part in project.parts:
        id_cell = ws.cell(row=row, column=PARTS_COL, value=part.id)
        id_cell.font = PART_ID_FONT
        id_cell.alignment = LEFT_CENTER
        id_cell.border = CELL_BORDER

        assigned = project.assigned_columns_for_part(part.id)

        for c_idx, col_key in enumerate(columns, start=PARTS_COL + 1):
            cell = ws.cell(row=row, column=c_idx)
            if col_key in assigned:
                cell.value = CHECK_MARK
                cell.font = CHECK_FONT
            cell.alignment = CENTER
            cell.border = CELL_BORDER

        row += 1

    last_row = row - 1

    # --- Column widths ---
    id_width = max([len("Parts")] + [len(p.id) for p in project.parts]) + 4
    ws.column_dimensions[get_column_letter(PARTS_COL)].width = id_width

    for c_idx, col_key in enumerate(columns, start=PARTS_COL + 1):
        width = max(len(col_key.variant), 8) + 4
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    ws.row_dimensions[HEADER_ROW_REGION].height = 22
    ws.row_dimensions[HEADER_ROW_VARIANT].height = 20

    # --- Frozen panes: keep Parts column + both header rows visible ---
    ws.freeze_panes = ws.cell(row=FIRST_DATA_ROW, column=PARTS_COL + 1)

    # --- Auto filter on the variant header row (row 2) across full range ---
    if last_row >= FIRST_DATA_ROW and last_col >= PARTS_COL:
        filter_range = (
            f"{get_column_letter(PARTS_COL)}{HEADER_ROW_VARIANT}:"
            f"{get_column_letter(last_col)}{last_row}"
        )
        ws.auto_filter.ref = filter_range

    wb.save(str(path))
